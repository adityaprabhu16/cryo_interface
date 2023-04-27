
import datetime
import json
import os
import sys
import time
from typing import Dict, Tuple

from openpyxl import Workbook


# Path to folder to combine data for.
target_dir: str = sys.argv[1]

# Read the metadata from metadata.json.
metadata_path = os.path.join(target_dir, 'metadata.json')
with open(metadata_path, encoding='utf-8') as mdata_file:
    metadata: Dict = json.load(mdata_file)

# Dictionary storing an integer timestamp and the corresponding temperatures (t1 and t2) as floats.
temps: Dict[int, Tuple[float, float]] = {}

# Read the temperatures from the csv file.
csv_path = os.path.join(target_dir, 'temperatures.csv')
with open(csv_path, encoding='utf-8') as csv_file:
    for line in csv_file.readlines():
        ts, t1, t2 = [float(x) for x in line.split(',')]
        temps[int(ts)] = t1,t2

# Create a excel workbook object.
wb = Workbook()

# Access the active worksheet.
ws = wb.active

# Store metadata into the first sheet in the workbook.
for i,(key,value) in enumerate(metadata.items()):
    row = i + 1
    ws[f'A{row}'] = key
    ws[f'B{row}'] = value

# Variables to keep track of the vna index.
vna1_count = 0
vna2_count = 0

# Iterate through data files in directory.
for file in sorted(os.listdir(target_dir)):

    # Check if the file is a .csv (and not the temperatures.csv file)
    if not file.endswith('.csv') or file == 'temperatures.csv':
        continue

    x = file.removesuffix('.csv').split('_')

    # Create integer timestamp from the file name.
    try:
        ts = int(time.mktime(datetime.datetime(int(x[0]), int(x[1]), int(x[2]), int(x[3]), int(x[4]), int(x[5])).timetuple()))
    except (IndexError, ValueError):
        print(f'WARNING: Encountered a .csv file with an invalid name: {file}')
        continue

    vna = x[-1]

    # Build name for the new sheet.
    if vna == 'vna1':
        vna1_count += 1
        sheet_name = f'v1_{vna1_count}'
        temp_sensor = metadata.get('vna1_temp')
    elif vna == 'vna2':
        vna2_count += 1
        sheet_name = f'v2_{vna2_count}'
        temp_sensor = metadata.get('vna2_temp')
    else:
        print(f'WARNING: Encountered a .csv file with an invalid name: {file}')
        continue
    
    # Create new sheet.
    wb.create_sheet(sheet_name)
    # Access the new sheet.
    ws = wb[sheet_name]

    # Add temperature at the top.
    ws['A1'] = 'Temperature (degC):'

    # Find the temperature that corresponds to the current timestamp.
    temp: Tuple[float, float] = temps.get(ts)

    # Check if temperature readings actually existed at this time.
    if temp is None:
        print(f'WARNING: Unable to find corresponding temperature for {sheet_name}.')
    else:
        if temp_sensor == 'temp1':
            ws['B1'] = temp[0]
        elif temp_sensor == 'temp2':
            ws['B1'] = temp[1]
        else:
            ws['B1'] = None

    # Add CSV below temperature.
    fpath = os.path.join(target_dir, file)
    with open(fpath, encoding='utf-8') as dataf:
        for lino, line in enumerate(dataf.readlines()):
            if ',' in line:
                a,b = line.split(',')
                ws[f'A{3+lino}'] = a
                ws[f'B{3+lino}'] = b
            else:
                ws[f'A{3+lino}'] = line


# Build path to save to.
wb_path = os.path.join(target_dir, 'combined_data.xlsx')

# Save the workbook.
wb.save(wb_path)

# Let the user know that we're done.
print('Done!')
